from DatabaseService.core import DatabaseService1, User, Filologiya, TTJ5, TTJ7, TTJ8, TTJ9, TTJ10
from DataConfig.file_path import get_file_path
from openpyxl import load_workbook, Workbook
from LoggingService import LoggerService
import asyncio

db = DatabaseService1(logger=LoggerService())

async def write_db():
    file_path = await get_file_path("Talabalar.xlsx")
    workbook = load_workbook(file_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    updated, skipped = 0, 0
    for row in rows:
        student_id = str(row[0]).strip() if row[0] else None
        jshshir = str(row[10]).strip() if row[10] else None
        print(student_id, jshshir)
        if not student_id or not jshshir:
            skipped += 1
            continue
        try:
            result = await db.update_by_field(
                User,
                field_name="username",
                field_value=student_id,
                updates={"jshshir": jshshir}
            )
            if result:
                updated += 1
            else:
                skipped += 1
        except Exception as ex:
            skipped += 1
    print(f"✅ Yangilangan: {updated}, ❌ O‘tkazib yuborilgan: {skipped}")


async def read_excel():
    file_path = await get_file_path("students.xlsx")
    workbook = load_workbook(file_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    try:
        for row in rows:
            print(row, end="\n")
            print(row[0])
            users = await db.get(User, filters={"jshshir": row[2]})
            await db.add(
                TTJ5(username=str(row[1]), password=str(row[2]), external_id=str(row[0]), full_name=str(row[4]) or 'None',
                     gender='Erkak',
                     faculty_id=int(row[6]) or 0, turniket_id=str(row[7]) or 'None', begin_time=(row[8]), end_time=(row[9]),
                     image_url='https://hemis.tdpu.uz/static//pi/b/4/b46a147cd59b84df2eafb2aceee64724.jpg',
                     created_at=row[11], student_id=str(row[12]), file_path=str(row[13]),
                     group_id=row[14] or 0, group_name=str(row[15]), speciality_id='None', speciality_name=str(row[17])
                     ))
            # break
    # (38, 1 355241102426, 2 'AD4373857', 3 None, 4 'PARDAYEVA MADINA MUSURMON QIZI', 5 'Ayol', 6 1, 7 20250001, 8 datetime.datetime(2025, 9, 6, 16, 15, 34, 78000), 9
    # 9 datetime.datetime(2029, 9, 5, 16, 15, 34, 78000), 10 'https://hemis.tdpu.uz/static/crop/2/6/320_320_90_2675932958.jpg', 11 datetime.datetime(2025, 9, 6, 16, 15, 34, 83000),
    # 12 355241102426, 13 'students\\355241102426_PARDAYEVA_MADINA_MUSURMON_QIZI_20250001.json', 14 1954, 15 'M uz 101 kechki 2025', 16 '8caf66ec-98a3-4f65-8582-0cf6ce38aa05', 17 'Matematika')
    except Exception as ex:
        print(ex)
    finally:
        workbook.close()
async def export_to_ttj5():
    file_path = await get_file_path("ttj_10.xlsx")
    workbook = load_workbook(file_path)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # 1-qator sarlavha

    added, skipped = 0, 0
    for row in rows:
        student_id = str(row[2]).strip() if row[0] else None
        print(student_id)
        if not student_id:
            skipped += 1
            continue

        # Users jadvalidan student_id bo‘yicha qidirish
        users = await db.get(User, filters={"jshshir": str(student_id)})
        if not users:
            # logger.warning(f"⚠️ User topilmadi: {student_id}")
            skipped += 1
            continue

        user = users[0]  # birinchi topilganini olish

        # Yangi TTJ5 yozuvini yaratish
        try:
            await db.add(
                TTJ10(
                    username=user.username,
                    password=user.password,
                    external_id=user.external_id,
                    student_id=user.student_id,
                    full_name=user.full_name or "None",
                    gender=user.gender or "Erkak",
                    faculty_id=user.faculty_id or 0,
                    turniket_id=user.turniket_id or "None",
                    begin_time=user.begin_time,
                    end_time=user.end_time,
                    image_url=user.image_url or "",
                    created_at=user.created_at,
                    file_path=user.file_path or "",
                    group_id=user.group_id or 0,
                    group_name=user.group_name or "None",
                    speciality_id=user.speciality_id or "None",
                    speciality_name=user.speciality_name or "None",
                    jshshir=user.jshshir or None
                )
            )
            added += 1
            # logger.info(f"✅ TTJ5 ga qo‘shildi: {student_id} | {user.full_name}")
        except Exception as ex:
            skipped += 1
            # logger.error(f"❌ Xato {student_id}: {ex}")

    print(f"✅ TTJ7 ga qo‘shilganlar: {added}, o'tkazib yuborilganlar: {skipped}")


if __name__ == "__main__":
    asyncio.run(export_to_ttj5())
